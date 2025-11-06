from typing import Dict, List, Tuple, Optional
import pandas as pd
from store_location import CheckAdress
from config.name import FilePath, TargetShipping, CompanyName, ExcelFieldName
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import json
from get_excel import GetExcelData


class ProductConfig:
    def __init__(self):
        with open("config/product_config.json", "r", encoding="utf-8") as f:
            self.config: Dict = json.load(f)

    def search_product(self, search_value: str, search_type: str = "mixx_name", c2c_name: str = None) -> str:
        for product_code, product_info in self.config.items():
            if search_type == "mixx_name" and search_value in product_info.get("mixx_name", []):
                return product_code
            elif search_type == "c2c_code" and c2c_name in product_info.get("c2c_name", []):
                c2c_code_list = product_info.get("c2c_code", [])
                for code in c2c_code_list:
                    if search_value in code:
                        return product_code
        return None


class OrderDataHandler:
    def __init__(self, platform: str):
        self.field_config = ExcelFieldName.get_config(platform)
        self.platform = platform
        self.product_config = ProductConfig()

    @staticmethod
    def format_date(order_time: str) -> str:
        try:
            if isinstance(order_time, datetime):
                date_time_format = order_time
            else:
                date_time_format = datetime.strptime(order_time, "%Y-%m-%d %H:%M:%S")
            return date_time_format.strftime("%Y%m%d")
        except (ValueError, TypeError):
            return "INVALID_DATE"

    def get_field_value(self, row: pd.Series, field_name: str) -> str:
        return str(row[self.field_config[field_name]])

    def get_product_code(self, row: pd.Series) -> str:
        if self.platform == "mixx":
            product_name = str(row["品名/規格"]).split("｜")[1] if "｜" in str(row["品名/規格"]) else str(row["品名/規格"])
            return self.product_config.search_product(search_type="mixx_name", search_value=product_name)
        elif self.platform == "c2c":
            return self.product_config.search_product(search_type="c2c_code", search_value=str(row["商品編號"]), c2c_name=str(row["商品樣式"]))
        elif self.platform == "shopline":
            return str(row[self.field_config["product_code"]])
        return ""

    def get_order_mark(self, row: pd.Series) -> str:
        platform_marks = {"c2c": ("減醣市集 X 快電商 C2C BUY", "出貨備註", " | "), "shopline": ("減醣市集", "出貨備註", "/"), "mixx": ("減醣市集", "備註", "/")}

        if self.platform in platform_marks:
            prefix, note_field, separator = platform_marks[self.platform]
            note = str(row[note_field])
            return f"{prefix}{'' if note == 'nan' else f'{separator}{note}'}"
        return ""

    def get_formatted_date(self, row: pd.Series) -> str:
        date_mapping = {
            "c2c": lambda r: self.format_date(r["建立時間"]),
            "shopline": lambda r: self.format_date(r["訂單日期"]),
            "mixx": lambda _: self.format_date(datetime.now()),
        }
        return date_mapping.get(self.platform, lambda _: "")(row)


class OrderProcessor:
    def __init__(self, platform: str):
        self.platform = platform
        self.data_handler = OrderDataHandler(platform)
        self.product_config = ProductConfig()

    def calculate_box_type(self, orders: List[Dict]) -> Tuple[str, str]:
        try:

            grand_total = 0
            for order in orders:
                if str(order["商品編號"]) and str(order["商品編號"]) != "nan":
                    qty = self.product_config.config[order["商品編號"]]["qty"]
                    order_quantity = int(float(order["訂購數量"]))
                    grand_total += qty * order_quantity

            if grand_total <= 14:
                return "box60-EA", "60公分紙箱"
            elif grand_total <= 47:
                return "box90-EA", "90公分紙箱"
            return "ERROR-需拆單", "ERROR-需拆單"
        except ValueError as e:
            print(f"處理商品編號時發生錯誤: {e}")
            raise

    def get_delivery_info(self, row: pd.Series, store_address: Dict) -> Tuple[str, str]:
        delivery_method = row["送貨方式"].split("（", 1)[0]

        if delivery_method == TargetShipping.tacat:
            return "Tcat", row["完整地址"]
        elif delivery_method == TargetShipping.family:
            store_name = row["門市名稱"]
            if store_name not in store_address[CompanyName.family]:
                return "全家", "ERROR"
            return "全家", f"{store_name} ({store_address[CompanyName.family][store_name]})"
        elif delivery_method == TargetShipping.seven:
            store_name = row["門市名稱"]
            if store_name not in store_address[CompanyName.seven]:
                return "7-11", "ERROR"
            return "7-11", f"(宅轉店){store_address[CompanyName.seven][store_name]}"
        return "UNKNOWN", "ERROR"

    def create_order_row(self, row: pd.Series, delivery_method: str = "Tcat") -> Dict:
        return {
            "貨主編號": "A442",
            "貨主單號\n(不同客戶端、不同溫層要分單)": self.data_handler.get_field_value(row, "order_id"),
            "客戶端代號(店號)": self.data_handler.get_field_value(row, "receiver_name"),
            "訂購日期": self.data_handler.get_formatted_date(row),
            "商品編號": self.data_handler.get_product_code(row),
            "商品名稱": self.data_handler.get_field_value(row, "product_name").replace("-F", ""),
            "訂購數量": self.data_handler.get_field_value(row, "product_quantity"),
            "配送方式\nFT : 逢泰配送\nTcat : 黑貓宅配\n全家到府取貨": delivery_method,
            "收貨人姓名": self.data_handler.get_field_value(row, "receiver_name"),
            "收貨人地址": row["address"] if self.platform == "shopline" else self.data_handler.get_field_value(row, "receiver_address"),
            "收貨人聯絡電話": self.data_handler.get_field_value(row, "receiver_phone"),
            "訂單 / 宅配單備註": self.data_handler.get_order_mark(row),
            "指定配送溫層\n001：常溫\n002：冷藏\n003：冷凍": "003",
            "品項備註": "",
        }

    def process_orders(self, data: pd.DataFrame, store_address: Optional[Dict] = None) -> List[Dict]:
        if self.platform == "shopline":
            return self._process_shopline_orders(data, store_address)
        elif self.platform == "mixx":
            return self._process_mixx_orders(data)
        elif self.platform == "c2c":
            return self._process_c2c_orders(data)
        return []

    def _add_box_to_order(self, personal_order: List[Dict]) -> Dict:
        box_type, box_name = self.calculate_box_type(personal_order)
        box_row = personal_order[0].copy()
        box_row.update(
            {
                "商品編號": box_type,
                "商品名稱": box_name,
                "訂購數量": "1",
                "品項備註": "箱子",
            }
        )
        return box_row

    def _process_mixx_orders(self, data: pd.DataFrame) -> List[Dict]:
        new_rows = []
        personal_order = []

        for _, row in data.iterrows():
            if personal_order and personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] != str(row["*銷售單號"]):
                new_rows.append(self._add_box_to_order(personal_order))
                personal_order.clear()

            new_row = self.create_order_row(row)
            if str(row["*銷售單號"]) != "nan":
                if not personal_order or personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] == str(row["*銷售單號"]):
                    personal_order.append(new_row)
                else:
                    personal_order = [new_row]
                new_rows.append(new_row)

        if personal_order:
            new_rows.append(self._add_box_to_order(personal_order))

        return new_rows

    def _process_c2c_orders(self, data: pd.DataFrame) -> List[Dict]:
        new_rows = []
        personal_order = []

        for _, row in data.iterrows():
            if str(row["商品編號"]) == "F2500000044":
                for i in range(2):
                    if personal_order and personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] != str(row["平台訂單編號"]):
                        new_rows.append(self._add_box_to_order(personal_order))
                        personal_order.clear()

                    new_row = self.create_order_row(row)
                    new_row["商品名稱"] = str(row["商品樣式"]).replace("(贈品)-F", "").split("+")[i]

                    if str(row["平台訂單編號"]) != "nan":
                        if not personal_order or personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] == str(row["平台訂單編號"]):
                            personal_order.append(new_row)
                        else:
                            personal_order = [new_row]
                        new_rows.append(new_row)
            else:
                if personal_order and personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] != str(row["平台訂單編號"]):
                    new_rows.append(self._add_box_to_order(personal_order))
                    personal_order.clear()

                new_row = self.create_order_row(row)

                if str(row["平台訂單編號"]) != "nan":
                    if not personal_order or personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] == str(row["平台訂單編號"]):
                        personal_order.append(new_row)
                    else:
                        personal_order = [new_row]
                    new_rows.append(new_row)

        if personal_order:
            new_rows.append(self._add_box_to_order(personal_order))

        return new_rows

    def _process_shopline_orders(self, data: pd.DataFrame, store_address: Dict) -> List[Dict]:
        new_rows = []
        personal_order = []
        skip_order = 0

        for _, row in data.iterrows():
            delivery_method, address = self.get_delivery_info(row, store_address)
            product_mark = "" if len(str(row["商品貨號"]).split("-")) < 3 else "-" + str(row["商品貨號"]).split("-")[2]

            if personal_order and personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] != str(row["訂單號碼"]):
                new_rows.append(self._add_box_to_order(personal_order))
                personal_order.clear()

            row_data = row.copy()
            row_data["address"] = address
            new_row = self.create_order_row(row_data, delivery_method)

            new_row["商品名稱"] = f"{row['商品名稱']}{product_mark}" if pd.notna(row["商品名稱"]) else f"{row['選項']}{product_mark}"

            if str(row["商品貨號"]) != "nan":
                if not personal_order or personal_order[0]["貨主單號\n(不同客戶端、不同溫層要分單)"] == str(row["訂單號碼"]):
                    personal_order.append(new_row)
                else:
                    personal_order = [new_row]
                if len(row) > 17 and pd.notna(row.iloc[17]) and row.iloc[17] in ["下午到貨", "上午到貨"]:
                    new_row["到貨時段\n1: 13點前\n2: 14~18\n3: 不限時"] = 1 if str(row.iloc[17]) == "上午到貨" else 2
                new_rows.append(new_row)
            else:
                skip_order += 1

        if personal_order:
            new_rows.append(self._add_box_to_order(personal_order))

        if skip_order > 0:
            print(f"\n商品貨號空白故扣除筆數：{skip_order}")

        return new_rows


class ReportGenerator:
    def __init__(self):
        self.product_config = ProductConfig()

    def apply_cell_format(self, cell):
        is_error = isinstance(cell.value, str) and "ERROR" in cell.value
        is_nan = isinstance(cell.value, str) and cell.value.lower() == "nan"

        cell.font = Font(name="微軟正黑體", size=11, bold=is_nan, color="FF0000" if (is_error or is_nan) else "000000")

        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def auto_adjust_dimensions(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column[1:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        for row in sheet.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    required_height = lines * 15
                    max_height = max(max_height, required_height)
            if max_height > 0:
                sheet.row_dimensions[cell.row].height = max_height

    def save_to_excel(self, rows: List[Dict], output_path: str):
        wb = load_workbook(FilePath.report)
        sheet = wb.active
        template_columns = [cell.value for cell in sheet[1]]

        processed_rows = [{col: row.get(col, "") for col in template_columns} for row in rows]
        df = pd.DataFrame(processed_rows)

        for i, row in enumerate(df.values, start=2):
            for j, value in enumerate(row, start=1):
                cell = sheet.cell(row=i, column=j, value=value)
                self.apply_cell_format(cell)

        self.auto_adjust_dimensions(sheet)
        wb.save(output_path)

    def generate_report(self, input_data_path: str, output_path: str, platform: str):
        if "." not in output_path:
            output_path += ".xlsx"

        excel = GetExcelData(input_data_path, platform)
        data = excel()
        field_config = ExcelFieldName.get_config(platform)

        order_processor = OrderProcessor(platform)

        if platform == "shopline":
            address_checker = CheckAdress(original_data_path=input_data_path)
            location_info = address_checker.check_adress()
            rows = order_processor.process_orders(data, location_info)
        else:
            rows = order_processor.process_orders(data)

        if platform == "c2c":
            product_code = field_config["product_code"]
            main_product_count = len(data[data[product_code] != "F2500000044"])
            freebies_count = len(data[data[product_code] == "F2500000044"])
            print(f"主商品數量: {main_product_count}\n")
            print(f"贈品數量: {freebies_count}\n")

        print(f"最終筆數: {len(rows)}\n")
        print(f"總訂單數: {len(data[field_config['order_id']].unique())}\n")

        self.save_to_excel(rows, output_path)


if __name__ == "__main__":
    generator = ReportGenerator()
    generator.generate_report(input_data_path=r"C:\Users\07711.Jason.Sung\OneDrive - Global ICT\文件\1106.xls", output_path="123", platform="shopline")
